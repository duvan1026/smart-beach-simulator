"""
Genera diagramas de analisis en nuevas hojas del Excel historico de Cadiz.
Fuente: Historico_Temp_Cadiz.xlsx — estacion AEMET 5973, Cadiz, 2024-2026.

Hojas generadas:
  - Analisis_Temperatura   : evolucion mensual tmin/tmax + boxplot por mes
  - Analisis_Viento        : evolucion racha diaria + distribucion acumulada + dias por umbral
  - Analisis_Presion       : evolucion presMax/presMin + boxplot por mes
  - Analisis_Escenarios    : frecuencia real de cada escenario segun umbrales Meteoalerta
"""

import io
import os
import statistics
from collections import defaultdict

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import numpy as np
import openpyxl
from openpyxl.drawing.image import Image as XLImage

# Obtiene la ruta absoluta de la carpeta donde está este script (la carpeta 'docs')
DIRECTORIO_BASE = os.path.dirname(os.path.abspath(__file__))

EXCEL_PATH = os.path.join(DIRECTORIO_BASE, "Historico_Temp_Cadiz.xlsx")
MESES_ES = ["", "Ene", "Feb", "Mar", "Abr", "May", "Jun",
            "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

# ── Colores por escenario ──────────────────────────────────────────────────────
COLOR_NORMAL    = "#4CAF50"
COLOR_LEVANTE   = "#FFC107"
COLOR_STORM     = "#FF9800"
COLOR_SEVERE    = "#F44336"
COLOR_HEAT      = "#E91E63"
COLOR_OCUPACION = "#9C27B0"


def parse_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "."))
    except ValueError:
        return None


def fig_to_image(fig, dpi=120):
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=dpi, bbox_inches="tight")
    buf.seek(0)
    plt.close(fig)
    return buf


def insert_image(ws, buf, cell, width_px=None, height_px=None):
    img = XLImage(buf)
    if width_px:
        img.width = width_px
    if height_px:
        img.height = height_px
    ws.add_image(img, cell)


def load_data(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    data = []
    for r in rows:
        fecha = str(r[0]) if r[0] else ""
        if len(fecha) < 7 or not fecha.startswith("20"):
            continue
        mes = int(fecha[5:7])
        anio = int(fecha[:4])
        tmin  = parse_float(r[7])
        tmax  = parse_float(r[9])
        vel   = parse_float(r[12])
        racha = parse_float(r[13])
        pmax  = parse_float(r[15])
        pmin  = parse_float(r[17])
        data.append({
            "fecha": fecha, "mes": mes, "anio": anio,
            "tmin": tmin, "tmax": tmax,
            "vel_kmh":   vel   * 3.6 if vel   is not None else None,
            "racha_kmh": racha * 3.6 if racha is not None else None,
            "pmax": pmax, "pmin": pmin,
        })
    return wb, data


def classify_scenario(d):
    """Clasifica un dia en el escenario mas severo segun umbrales Meteoalerta costeros."""
    racha = d["racha_kmh"]
    tmax  = d["tmax"]
    if racha is None:
        return "NORMAL"
    if racha >= 90:
        return "SEVERE_STORM"
    if racha >= 60:
        return "STORM"
    if tmax is not None and tmax >= 36:
        return "HEATWAVE"
    if racha >= 50:
        return "LEVANTE"
    return "NORMAL"


# ── HOJA 1: Temperatura ────────────────────────────────────────────────────────
def sheet_temperatura(wb, data):
    ws = wb.create_sheet("Analisis_Temperatura")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "ANALISIS DE TEMPERATURA — Estacion 5973 Cadiz (AEMET)"
    ws["A1"].font = openpyxl.styles.Font(bold=True, size=13)
    ws["A2"] = "Fuente: Historico_Temp_Cadiz.xlsx | Umbrales: METEOALERTA_ANX1 zona 611103 Litoral gaditano"
    ws["A2"].font = openpyxl.styles.Font(italic=True, size=9, color="666666")

    fechas = [d["fecha"] for d in data if d["tmax"] is not None]
    tmax_s = [d["tmax"] for d in data if d["tmax"] is not None]
    tmin_s = [d["tmin"] for d in data if d["tmin"] is not None and d["tmax"] is not None]

    # ── Grafico 1: Evolucion diaria tmin/tmax con bandas de alerta ──
    fig, ax = plt.subplots(figsize=(16, 4))
    ax.fill_between(range(len(fechas)), tmin_s, tmax_s, alpha=0.25, color="#2196F3", label="Rango diario tmin-tmax")
    ax.plot(range(len(fechas)), tmax_s, color="#F44336", linewidth=0.8, label="Tmax")
    ax.plot(range(len(fechas)), tmin_s, color="#2196F3", linewidth=0.8, label="Tmin")

    # Bandas de alerta de temperatura (AEMET zona 611103)
    ax.axhline(36, color="#FFC107", linewidth=1.2, linestyle="--", label="Umbral amarillo (36°C)")
    ax.axhline(39, color="#FF9800", linewidth=1.2, linestyle="--", label="Umbral naranja (39°C)")
    ax.axhline(42, color="#F44336", linewidth=1.2, linestyle="--", label="Umbral rojo (42°C)")

    # Eje X con etiquetas de mes/año
    step = max(1, len(fechas) // 20)
    xticks = list(range(0, len(fechas), step))
    ax.set_xticks(xticks)
    ax.set_xticklabels([fechas[i][:7] for i in xticks], rotation=45, ha="right", fontsize=7)
    ax.set_ylabel("Temperatura (°C)")
    ax.set_title("Evolucion diaria de temperatura — Cadiz 2024-2026", fontweight="bold")
    ax.legend(fontsize=7, ncol=3, loc="upper left")
    ax.grid(axis="y", alpha=0.3)
    fig.tight_layout()
    buf1 = fig_to_image(fig)

    # ── Grafico 2: Boxplot tmax por mes ──
    monthly_tmax = defaultdict(list)
    monthly_tmin = defaultdict(list)
    for d in data:
        if d["tmax"] is not None:
            monthly_tmax[d["mes"]].append(d["tmax"])
        if d["tmin"] is not None:
            monthly_tmin[d["mes"]].append(d["tmin"])

    meses = sorted(monthly_tmax.keys())
    fig, axes = plt.subplots(1, 2, figsize=(14, 5))

    # Boxplot tmax
    bp = axes[0].boxplot(
        [monthly_tmax[m] for m in meses],
        labels=[MESES_ES[m] for m in meses],
        patch_artist=True, notch=False, showfliers=True
    )
    colors_box = ["#FFCDD2" if statistics.mean(monthly_tmax[m]) >= 36 else "#BBDEFB" for m in meses]
    for patch, color in zip(bp["boxes"], colors_box):
        patch.set_facecolor(color)
    axes[0].axhline(36, color="#FFC107", linewidth=1.2, linestyle="--", label="Umbral amarillo 36°C")
    axes[0].axhline(39, color="#FF9800", linewidth=1.2, linestyle="--", label="Umbral naranja 39°C")
    axes[0].set_title("Distribucion mensual — Temperatura maxima", fontweight="bold")
    axes[0].set_ylabel("Tmax (°C)")
    axes[0].legend(fontsize=8)
    axes[0].grid(axis="y", alpha=0.3)

    # Media mensual tmin vs tmax
    medias_tmax = [statistics.mean(monthly_tmax[m]) for m in meses]
    medias_tmin = [statistics.mean(monthly_tmin[m]) for m in meses]
    x = np.arange(len(meses))
    axes[1].bar(x - 0.2, medias_tmax, 0.4, label="Media Tmax", color="#EF9A9A")
    axes[1].bar(x + 0.2, medias_tmin, 0.4, label="Media Tmin", color="#90CAF9")
    axes[1].set_xticks(x)
    axes[1].set_xticklabels([MESES_ES[m] for m in meses])
    axes[1].set_title("Media mensual Tmin / Tmax", fontweight="bold")
    axes[1].set_ylabel("Temperatura (°C)")
    axes[1].legend(fontsize=8)
    axes[1].grid(axis="y", alpha=0.3)
    fig.tight_layout()
    buf2 = fig_to_image(fig)

    insert_image(ws, buf1, "A4",  width_px=1100, height_px=280)
    insert_image(ws, buf2, "A22", width_px=1100, height_px=380)

    # Tabla de estadisticos por mes
    ws["A57"] = "Estadisticos mensuales de temperatura (°C)"
    ws["A57"].font = openpyxl.styles.Font(bold=True)
    headers = ["Mes", "Tmin_mean", "Tmin_min", "Tmax_mean", "Tmax_max", "Dias_alerta_amarilla(≥36)", "Dias_alerta_naranja(≥39)", "Dias_alerta_roja(≥42)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=58, column=col, value=h)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="D0D0D0")
    for row_i, m in enumerate(meses, 59):
        tmax_list = monthly_tmax[m]
        tmin_list = monthly_tmin[m]
        ws.cell(row=row_i, column=1, value=MESES_ES[m])
        ws.cell(row=row_i, column=2, value=round(statistics.mean(tmin_list), 1))
        ws.cell(row=row_i, column=3, value=round(min(tmin_list), 1))
        ws.cell(row=row_i, column=4, value=round(statistics.mean(tmax_list), 1))
        ws.cell(row=row_i, column=5, value=round(max(tmax_list), 1))
        ws.cell(row=row_i, column=6, value=sum(1 for v in tmax_list if v >= 36))
        ws.cell(row=row_i, column=7, value=sum(1 for v in tmax_list if v >= 39))
        ws.cell(row=row_i, column=8, value=sum(1 for v in tmax_list if v >= 42))

    print("  Hoja Analisis_Temperatura creada")


# ── HOJA 2: Viento ─────────────────────────────────────────────────────────────
def sheet_viento(wb, data):
    ws = wb.create_sheet("Analisis_Viento")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "ANALISIS DE VIENTO — Estacion 5973 Cadiz (AEMET)"
    ws["A1"].font = openpyxl.styles.Font(bold=True, size=13)
    ws["A2"] = "Fuente: Historico_Temp_Cadiz.xlsx | Umbrales: METEOALERTA_ANX1 sec. 2.2 Fenomenos costeros atlanticos"
    ws["A2"].font = openpyxl.styles.Font(italic=True, size=9, color="666666")

    fechas  = [d["fecha"]    for d in data if d["racha_kmh"] is not None]
    rachas  = [d["racha_kmh"] for d in data if d["racha_kmh"] is not None]
    vel_med = [d["vel_kmh"]  for d in data if d["vel_kmh"]  is not None]

    # ── Grafico 1: Evolucion diaria racha con bandas de alerta ──
    fig, ax = plt.subplots(figsize=(16, 4))
    ax.bar(range(len(rachas)), rachas, color=[
        COLOR_SEVERE if v >= 90 else
        COLOR_STORM  if v >= 60 else
        COLOR_LEVANTE if v >= 50 else
        "#90CAF9"
        for v in rachas
    ], width=1.0, label="Racha maxima diaria")
    ax.plot(range(len(vel_med)), vel_med, color="#0D47A1", linewidth=0.7, label="Vel. media diaria")

    ax.axhline(50, color=COLOR_LEVANTE, linewidth=1.3, linestyle="--", label="Umbral amarillo 50 km/h")
    ax.axhline(60, color=COLOR_STORM,   linewidth=1.3, linestyle="--", label="Umbral naranja 60 km/h")
    ax.axhline(90, color=COLOR_SEVERE,  linewidth=1.3, linestyle="--", label="Umbral rojo 90 km/h")

    step = max(1, len(fechas) // 20)
    xticks = list(range(0, len(fechas), step))
    ax.set_xticks(xticks)
    ax.set_xticklabels([fechas[i][:7] for i in xticks], rotation=45, ha="right", fontsize=7)
    ax.set_ylabel("Viento (km/h)")
    ax.set_title("Evolucion diaria de viento — Cadiz 2024-2026\n(amarillo≥50, naranja≥60, rojo≥90 km/h — AEMET sec.2.2 costero)", fontweight="bold")
    ax.legend(fontsize=7, ncol=3)
    ax.grid(axis="y", alpha=0.3)
    fig.tight_layout()
    buf1 = fig_to_image(fig)

    # ── Grafico 2: Distribucion acumulada y frecuencia por umbral ──
    fig, axes = plt.subplots(1, 2, figsize=(14, 5))

    # CDF de racha
    sorted_r = sorted(rachas)
    cdf = np.arange(1, len(sorted_r) + 1) / len(sorted_r) * 100
    axes[0].plot(sorted_r, cdf, color="#1565C0", linewidth=2)
    for umbral, color, label in [(50, COLOR_LEVANTE, "50 km/h amarillo"), (60, COLOR_STORM, "60 km/h naranja"), (90, COLOR_SEVERE, "90 km/h rojo")]:
        pct_val = sum(1 for v in rachas if v < umbral) / len(rachas) * 100
        axes[0].axvline(umbral, color=color, linewidth=1.3, linestyle="--")
        axes[0].annotate(f"{label}\n{pct_val:.0f}% dias < umbral", xy=(umbral, pct_val),
                         xytext=(umbral + 2, pct_val - 8), fontsize=7,
                         arrowprops=dict(arrowstyle="->", color=color), color=color)
    axes[0].set_xlabel("Racha maxima (km/h)")
    axes[0].set_ylabel("% dias con racha < valor")
    axes[0].set_title("Distribucion acumulada de racha maxima", fontweight="bold")
    axes[0].grid(alpha=0.3)

    # Dias por umbral por mes
    monthly_racha = defaultdict(list)
    for d in data:
        if d["racha_kmh"] is not None:
            monthly_racha[d["mes"]].append(d["racha_kmh"])
    meses = sorted(monthly_racha.keys())
    dias_50 = [sum(1 for v in monthly_racha[m] if v >= 50) for m in meses]
    dias_60 = [sum(1 for v in monthly_racha[m] if v >= 60) for m in meses]
    dias_90 = [sum(1 for v in monthly_racha[m] if v >= 90) for m in meses]
    x = np.arange(len(meses))
    w = 0.25
    axes[1].bar(x - w, dias_50, w, label="≥50 km/h (amarillo)", color=COLOR_LEVANTE)
    axes[1].bar(x,     dias_60, w, label="≥60 km/h (naranja)",  color=COLOR_STORM)
    axes[1].bar(x + w, dias_90, w, label="≥90 km/h (rojo)",     color=COLOR_SEVERE)
    axes[1].set_xticks(x)
    axes[1].set_xticklabels([MESES_ES[m] for m in meses])
    axes[1].set_title("Dias por umbral de alerta de viento por mes", fontweight="bold")
    axes[1].set_ylabel("Numero de dias")
    axes[1].legend(fontsize=8)
    axes[1].grid(axis="y", alpha=0.3)
    fig.tight_layout()
    buf2 = fig_to_image(fig)

    # ── Grafico 3: Boxplot racha por mes ──
    fig, ax = plt.subplots(figsize=(14, 4))
    bp = ax.boxplot(
        [monthly_racha[m] for m in meses],
        labels=[MESES_ES[m] for m in meses],
        patch_artist=True, showfliers=True
    )
    for patch in bp["boxes"]:
        patch.set_facecolor("#B3E5FC")
    ax.axhline(50, color=COLOR_LEVANTE, linewidth=1.2, linestyle="--", label="50 km/h amarillo")
    ax.axhline(60, color=COLOR_STORM,   linewidth=1.2, linestyle="--", label="60 km/h naranja")
    ax.axhline(90, color=COLOR_SEVERE,  linewidth=1.2, linestyle="--", label="90 km/h rojo")
    ax.set_title("Distribucion mensual de racha maxima de viento", fontweight="bold")
    ax.set_ylabel("Racha maxima (km/h)")
    ax.legend(fontsize=8)
    ax.grid(axis="y", alpha=0.3)
    fig.tight_layout()
    buf3 = fig_to_image(fig)

    insert_image(ws, buf1, "A4",  width_px=1100, height_px=280)
    insert_image(ws, buf2, "A22", width_px=1100, height_px=370)
    insert_image(ws, buf3, "A42", width_px=1100, height_px=290)

    # Tabla estadisticos
    ws["A62"] = "Estadisticos mensuales de viento (km/h)"
    ws["A62"].font = openpyxl.styles.Font(bold=True)
    headers = ["Mes", "Vel_media_mean", "Racha_mean", "Racha_max", "Dias≥50(amarillo)", "Dias≥60(naranja)", "Dias≥90(rojo)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=63, column=col, value=h)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="D0D0D0")
    monthly_vel = defaultdict(list)
    for d in data:
        if d["vel_kmh"] is not None:
            monthly_vel[d["mes"]].append(d["vel_kmh"])
    for row_i, m in enumerate(meses, 64):
        ws.cell(row=row_i, column=1, value=MESES_ES[m])
        ws.cell(row=row_i, column=2, value=round(statistics.mean(monthly_vel[m]), 1) if monthly_vel[m] else "")
        ws.cell(row=row_i, column=3, value=round(statistics.mean(monthly_racha[m]), 1))
        ws.cell(row=row_i, column=4, value=round(max(monthly_racha[m]), 1))
        ws.cell(row=row_i, column=5, value=sum(1 for v in monthly_racha[m] if v >= 50))
        ws.cell(row=row_i, column=6, value=sum(1 for v in monthly_racha[m] if v >= 60))
        ws.cell(row=row_i, column=7, value=sum(1 for v in monthly_racha[m] if v >= 90))

    print("  Hoja Analisis_Viento creada")


# ── HOJA 3: Presion ────────────────────────────────────────────────────────────
def sheet_presion(wb, data):
    ws = wb.create_sheet("Analisis_Presion")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "ANALISIS DE PRESION ATMOSFERICA — Estacion 5973 Cadiz (AEMET)"
    ws["A1"].font = openpyxl.styles.Font(bold=True, size=13)
    ws["A2"] = "Fuente: Historico_Temp_Cadiz.xlsx | Presion minima historica: 994.8 hPa (AEMET-HIST)"
    ws["A2"].font = openpyxl.styles.Font(italic=True, size=9, color="666666")

    data_p = [d for d in data if d["pmax"] is not None and d["pmin"] is not None]
    fechas = [d["fecha"] for d in data_p]
    pmax_s = [d["pmax"]  for d in data_p]
    pmin_s = [d["pmin"]  for d in data_p]

    # ── Grafico 1: Evolucion diaria presMax/presMin ──
    fig, ax = plt.subplots(figsize=(16, 4))
    ax.fill_between(range(len(fechas)), pmin_s, pmax_s, alpha=0.2, color="#7B1FA2")
    ax.plot(range(len(fechas)), pmax_s, color="#7B1FA2", linewidth=0.8, label="Presion maxima diaria")
    ax.plot(range(len(fechas)), pmin_s, color="#E91E63", linewidth=0.8, label="Presion minima diaria")

    # Lineas de referencia para escenarios
    ax.axhline(1010, color=COLOR_LEVANTE, linewidth=1.1, linestyle="--", label="1010 hPa (limite LEVANTE)")
    ax.axhline(997,  color=COLOR_STORM,   linewidth=1.1, linestyle="--", label="997 hPa (limite STORM)")
    ax.axhline(994,  color=COLOR_SEVERE,  linewidth=1.1, linestyle="--", label="994 hPa (limite SEVERE — min historico)")

    step = max(1, len(fechas) // 20)
    xticks = list(range(0, len(fechas), step))
    ax.set_xticks(xticks)
    ax.set_xticklabels([fechas[i][:7] for i in xticks], rotation=45, ha="right", fontsize=7)
    ax.set_ylabel("Presion (hPa)")
    ax.set_title("Evolucion diaria de presion atmosferica — Cadiz 2024-2026", fontweight="bold")
    ax.legend(fontsize=7, ncol=3)
    ax.grid(axis="y", alpha=0.3)
    fig.tight_layout()
    buf1 = fig_to_image(fig)

    # ── Grafico 2: Boxplot presion por mes ──
    monthly_pmax = defaultdict(list)
    monthly_pmin = defaultdict(list)
    for d in data_p:
        monthly_pmax[d["mes"]].append(d["pmax"])
        monthly_pmin[d["mes"]].append(d["pmin"])
    meses = sorted(monthly_pmax.keys())

    fig, axes = plt.subplots(1, 2, figsize=(14, 5))
    bp1 = axes[0].boxplot([monthly_pmax[m] for m in meses],
                          labels=[MESES_ES[m] for m in meses], patch_artist=True)
    for p in bp1["boxes"]: p.set_facecolor("#E1BEE7")
    axes[0].set_title("Distribucion mensual — Presion maxima", fontweight="bold")
    axes[0].set_ylabel("Presion (hPa)")
    axes[0].grid(axis="y", alpha=0.3)

    bp2 = axes[1].boxplot([monthly_pmin[m] for m in meses],
                          labels=[MESES_ES[m] for m in meses], patch_artist=True)
    for p in bp2["boxes"]: p.set_facecolor("#F8BBD9")
    axes[1].axhline(997, color=COLOR_STORM,  linewidth=1.2, linestyle="--", label="997 hPa STORM")
    axes[1].axhline(994, color=COLOR_SEVERE, linewidth=1.2, linestyle="--", label="994 hPa SEVERE")
    axes[1].set_title("Distribucion mensual — Presion minima", fontweight="bold")
    axes[1].set_ylabel("Presion (hPa)")
    axes[1].legend(fontsize=8)
    axes[1].grid(axis="y", alpha=0.3)
    fig.tight_layout()
    buf2 = fig_to_image(fig)

    insert_image(ws, buf1, "A4",  width_px=1100, height_px=280)
    insert_image(ws, buf2, "A22", width_px=1100, height_px=370)

    # Tabla estadisticos
    ws["A44"] = "Estadisticos mensuales de presion (hPa)"
    ws["A44"].font = openpyxl.styles.Font(bold=True)
    headers = ["Mes", "Pmax_mean", "Pmax_max", "Pmin_mean", "Pmin_min", "Dias_pmin<997", "Dias_pmin<994"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=45, column=col, value=h)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="D0D0D0")
    for row_i, m in enumerate(meses, 46):
        ws.cell(row=row_i, column=1, value=MESES_ES[m])
        ws.cell(row=row_i, column=2, value=round(statistics.mean(monthly_pmax[m]), 1))
        ws.cell(row=row_i, column=3, value=round(max(monthly_pmax[m]), 1))
        ws.cell(row=row_i, column=4, value=round(statistics.mean(monthly_pmin[m]), 1))
        ws.cell(row=row_i, column=5, value=round(min(monthly_pmin[m]), 1))
        ws.cell(row=row_i, column=6, value=sum(1 for v in monthly_pmin[m] if v < 997))
        ws.cell(row=row_i, column=7, value=sum(1 for v in monthly_pmin[m] if v < 994))

    print("  Hoja Analisis_Presion creada")


# ── HOJA 4: Escenarios ─────────────────────────────────────────────────────────
def sheet_escenarios(wb, data):
    ws = wb.create_sheet("Analisis_Escenarios")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "FRECUENCIA DE ESCENARIOS — Clasificacion segun umbrales AEMET Meteoalerta costeros"
    ws["A1"].font = openpyxl.styles.Font(bold=True, size=13)
    ws["A2"] = ("Amarillo(Levante)≥50 km/h | Naranja(Storm)≥60 km/h | Rojo(Severe)≥90 km/h | "
                "Ola de calor: tmax≥36°C | Normal: resto")
    ws["A2"].font = openpyxl.styles.Font(italic=True, size=9, color="666666")

    escenarios = [classify_scenario(d) for d in data]
    ORDEN = ["NORMAL", "LEVANTE", "STORM", "SEVERE_STORM", "HEATWAVE"]
    COLORES = {
        "NORMAL":       COLOR_NORMAL,
        "LEVANTE":      COLOR_LEVANTE,
        "STORM":        COLOR_STORM,
        "SEVERE_STORM": COLOR_SEVERE,
        "HEATWAVE":     COLOR_HEAT,
    }
    ETIQUETAS = {
        "NORMAL":       "Normal",
        "LEVANTE":      "Levante (≥50 km/h)",
        "STORM":        "Tormenta (≥60 km/h)",
        "SEVERE_STORM": "Tormenta Severa (≥90 km/h)",
        "HEATWAVE":     "Ola de Calor (tmax≥36°C)",
    }

    conteo = {e: escenarios.count(e) for e in ORDEN}
    total  = len(escenarios)

    # # ── Grafico 1: Pie de frecuencia global ──
    # fig, axes = plt.subplots(1, 2, figsize=(14, 5))
    # sizes  = [conteo[e] for e in ORDEN if conteo[e] > 0]
    # labels = [f"{ETIQUETAS[e]}\n{conteo[e]} dias ({conteo[e]/total*100:.1f}%)" for e in ORDEN if conteo[e] > 0]
    # colors = [COLORES[e] for e in ORDEN if conteo[e] > 0]
    # axes[0].pie(sizes, labels=labels, colors=colors, autopct="", startangle=90,
    #             wedgeprops={"edgecolor": "white", "linewidth": 1.5})
    # axes[0].set_title(f"Frecuencia global de escenarios\n(total: {total} dias)", fontweight="bold")

    # ── Grafico 1: Pie de frecuencia global ──
    fig, axes = plt.subplots(1, 2, figsize=(14, 5))
    sizes  = [conteo[e] for e in ORDEN if conteo[e] > 0]
    
    # Cambiamos el salto de linea (\n) por dos puntos (:) para que quede en una sola linea
    labels = [f"{ETIQUETAS[e]}: {conteo[e]} dias ({conteo[e]/total*100:.1f}%)" for e in ORDEN if conteo[e] > 0]
    colors = [COLORES[e] for e in ORDEN if conteo[e] > 0]
    
    # Guardamos los elementos (wedges) y QUITAMOS 'labels=labels' para no amontonar texto
    wedges, texts = axes[0].pie(sizes, colors=colors, startangle=90,
                                wedgeprops={"edgecolor": "white", "linewidth": 1.5})
    
    axes[0].set_title(f"Frecuencia global de escenarios\n(total: {total} dias)", fontweight="bold")
    
    # Agregamos la leyenda debajo del grafico circular
    axes[0].legend(wedges, labels, loc="upper center", bbox_to_anchor=(0.5, -0.05), 
                   fontsize=9, ncol=1, frameon=False)

    # Barras de frecuencia anual (Este código se mantiene igual, lo pongo para que te guíes)
    anios = sorted(set(d["anio"] for d in data))

    # Barras de frecuencia anual
    anios = sorted(set(d["anio"] for d in data))
    x = np.arange(len(anios))
    w = 0.15
    for i, esc in enumerate(ORDEN):
        vals = [sum(1 for j, d in enumerate(data) if d["anio"] == a and escenarios[j] == esc) for a in anios]
        axes[1].bar(x + i * w, vals, w, label=ETIQUETAS[esc], color=COLORES[esc])
    axes[1].set_xticks(x + w * 2)
    axes[1].set_xticklabels([str(a) for a in anios])
    axes[1].set_title("Frecuencia de escenarios por ano", fontweight="bold")
    axes[1].set_ylabel("Numero de dias")
    axes[1].legend(fontsize=7)
    axes[1].grid(axis="y", alpha=0.3)
    fig.tight_layout()
    buf1 = fig_to_image(fig)

    # ── Grafico 2: Frecuencia mensual de escenarios (barras apiladas) ──
    monthly_esc = defaultdict(lambda: defaultdict(int))
    for d, esc in zip(data, escenarios):
        monthly_esc[d["mes"]][esc] += 1
    meses = sorted(monthly_esc.keys())

    fig, ax = plt.subplots(figsize=(14, 5))
    bottom = np.zeros(len(meses))
    for esc in ORDEN:
        vals = np.array([monthly_esc[m][esc] for m in meses], dtype=float)
        ax.bar([MESES_ES[m] for m in meses], vals, bottom=bottom,
               label=ETIQUETAS[esc], color=COLORES[esc])
        bottom += vals
    ax.set_title("Composicion mensual de escenarios — dias por nivel de alerta", fontweight="bold")
    ax.set_ylabel("Numero de dias")
    ax.legend(fontsize=8, loc="upper right")
    ax.grid(axis="y", alpha=0.3)
    fig.tight_layout()
    buf2 = fig_to_image(fig)

    # ── Grafico 3: Linea temporal coloreada por escenario ──
    fig, ax = plt.subplots(figsize=(16, 2.5))
    for i, (d, esc) in enumerate(zip(data, escenarios)):
        ax.bar(i, 1, color=COLORES[esc], width=1.0, linewidth=0)
    step = max(1, len(data) // 20)
    xticks = list(range(0, len(data), step))
    ax.set_xticks(xticks)
    ax.set_xticklabels([data[i]["fecha"][:7] for i in xticks], rotation=45, ha="right", fontsize=7)
    ax.set_yticks([])
    patches = [mpatches.Patch(color=COLORES[e], label=ETIQUETAS[e]) for e in ORDEN]
    ax.legend(handles=patches, fontsize=7, loc="upper right", ncol=3)
    ax.set_title("Linea temporal de escenarios — Cadiz 2024-2026", fontweight="bold")
    fig.tight_layout()
    buf3 = fig_to_image(fig)

    insert_image(ws, buf1, "A4",  width_px=1100, height_px=370)
    insert_image(ws, buf2, "A26", width_px=1100, height_px=370)
    insert_image(ws, buf3, "A48", width_px=1100, height_px=200)

    # Tabla resumen
    ws["A58"] = "Resumen de frecuencia de escenarios"
    ws["A58"].font = openpyxl.styles.Font(bold=True)
    headers = ["Escenario", "Criterio", "Dias", "Porcentaje (%)"]
    criterios = {
        "NORMAL":       "racha < 50 km/h y tmax < 36°C",
        "LEVANTE":      "racha 50–59 km/h",
        "STORM":        "racha 60–89 km/h",
        "SEVERE_STORM": "racha ≥ 90 km/h",
        "HEATWAVE":     "tmax ≥ 36°C (sin viento fuerte)",
    }
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=59, column=col, value=h)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="D0D0D0")
    for row_i, esc in enumerate(ORDEN, 60):
        ws.cell(row=row_i, column=1, value=ETIQUETAS[esc])
        ws.cell(row=row_i, column=2, value=criterios[esc])
        ws.cell(row=row_i, column=3, value=conteo[esc])
        ws.cell(row=row_i, column=4, value=round(conteo[esc] / total * 100, 1))

    print("  Hoja Analisis_Escenarios creada")


# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    print(f"Cargando datos de {EXCEL_PATH} ...")
    wb, data = load_data(EXCEL_PATH)
    print(f"  {len(data)} registros cargados")

    # Eliminar hojas de analisis previas si existen (para re-ejecutar el script)
    for nombre in ["Analisis_Temperatura", "Analisis_Viento", "Analisis_Presion", "Analisis_Escenarios"]:
        if nombre in wb.sheetnames:
            del wb[nombre]
            print(f"  Hoja '{nombre}' eliminada (se regenera)")

    print("Generando hojas ...")
    sheet_temperatura(wb, data)
    sheet_viento(wb, data)
    sheet_presion(wb, data)
    sheet_escenarios(wb, data)

    wb.save(EXCEL_PATH)
    print(f"\nGuardado en {EXCEL_PATH}")
    print("Hojas generadas: Analisis_Temperatura, Analisis_Viento, Analisis_Presion, Analisis_Escenarios")


if __name__ == "__main__":
    main()
